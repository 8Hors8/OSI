import openpyxl as op
import bank_xlsx

from time import time


class Assistant:
    """Класс помощник"""

    def __init__(self, path_ved: str, path_bank: str, kv_quantity: int):

        self.path_ved = path_ved
        self.path_bank = path_bank
        self.kv_quantity = kv_quantity  # колличесво квартир

        wb_ved = op.load_workbook(path_ved)  # открываем фаил с ведомостью
        wb_bank = op.load_workbook(path_bank, data_only=True)  # открываем фаил от банка
        self.wb_ved = wb_ved
        self.wb_bank = wb_bank

        self.name_ved = path_ved.split('/')[-1]
        self.name_bank = path_bank.split('/')[-1]

        self.ved_sheet_list = wb_ved['список как должн']  # открываем лист с ведомостью
        self.ved_sheet_payment = wb_ved['оплата']  # открываем лист с оплатой
        self.bank_sheet = wb_bank.active  # открываем активный лист в файле от банка

        self.bank_dict = None  # словарь квартир с суммой и датой оплаты
        self.mon_coord_dict = None  # словарь с числом месяца и координатами ячейки
        self.summ_bank = None  # сумма из файла банка
        self.payments_row = None
        self.payments_coord_colu = None
        self.payments_coord_row = None

    def bank_reading(self):
        """Получаем словарь с квартирами суммой и датами  """
        self.bank_dict = bank_xlsx.payments(self.bank_sheet, self.name_bank)
        # print('bank_reading', self.bank_dict)

    def month_coordin(self):  # получаем словарь с месяцем и координатами
        """Получаем словарь с числом месяц и координаты ячейки """
        maxrow_sheet_paym = self.ved_sheet_payment.max_row
        month_dict = {'01': 'январь',
                      '02': 'февраль',
                      '03': 'март',
                      '04': 'апрель',
                      '05': 'май',
                      '06': 'июнь',
                      '07': 'июль',
                      '08': 'август',
                      '09': 'сентябрь',
                      '10': 'октябрь',
                      '11': 'ноябрь',
                      '12': 'декабрь'
                      }
        inv_month_dict = {v: k for k, v in month_dict.items()}  # переворачиваем словарь
        # print(inv_month_dict.keys())

        self.mon_coord_dict = {}
        for row in range(1, maxrow_sheet_paym):
            value_sheet_paym = self.ved_sheet_payment.cell(row=row, column=1).value

            if value_sheet_paym in inv_month_dict.keys():
                key = inv_month_dict.get(value_sheet_paym)

                # print(f'month_coordin  {value_sheet_paym} {inv_month_dict.get(value_sheet_paym)} row {row} {1}')
                self.mon_coord_dict[key] = (row, 1)

        # print(self.mon_coord_dict)

    def record_payments(self):
        """Запись платежей"""
        re = []
        for t in range(1, self.kv_quantity + 1):
            re.append(str(t))

        for kv in self.bank_dict.keys():
            self.summ_bank = self.bank_dict[kv][0]
            date = self.bank_dict[kv][1]
            month = date.split('.')[1]

            kv = kv.split("-")[0]
            if kv in re:

                # print('record_payments',month)

                coord_row = self.mon_coord_dict[month][0]
                coord_colu = self.mon_coord_dict[month][1]
                self.payments_coord_row = coord_row
                self.payments_coord_colu = coord_colu
                # print('record_payments',coord_row,coord_colu)

                for i in range(2, self.kv_quantity + 2):

                    kv_payment = self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu + 1).value

                    if kv == str(kv_payment):
                        date_payment = self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu).value
                        date_payment = date_payment if date_payment is not None else str(0)
                        date_payment = str(date_payment)
                        # print('kv_payment',date_payment)

                        if date_payment == "0":
                            self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu, value=date)
                            self.record_payments_summ(coord_row, coord_colu, i)
                            self.record_ved(kv)

                            break

                        elif date not in date_payment.split('/'):

                            record_date = f'{str(date_payment)}/{date}'
                            self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu, value=record_date)
                            self.record_payments_summ(coord_row, coord_colu, i)
                            self.record_ved(kv)
                            break

                        else:
                            break

                    else:
                        pass

    def record_payments_summ(self, coord_row, coord_colu, i):
        """Записываем в листе оплата сумму из банковского файла """
        summ_pyment = self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu + 3).value
        summ_pyment = summ_pyment if summ_pyment is not None else 0
        summ_pyment_bank = summ_pyment + self.summ_bank
        self.ved_sheet_payment.cell(row=coord_row + i, column=coord_colu + 3,
                                    value=summ_pyment_bank)
        # print('kv_payment', type(date_payment),type(date))
        self.payments_row = coord_row + i

    def record_ved(self, kv):
        """Запись ведомости за должности"""
        for row in range(8, self.kv_quantity + 8):
            kv_ved = str(self.ved_sheet_list.cell(row=row, column=2).value)
            if kv == kv_ved:
                summ_ved_fee = self.ved_sheet_list.cell(row=row, column=6).value  # ежемесячный взнос
                summ_ved_duty = self.ved_sheet_list.cell(row=row, column=8).value  # долг за прошлый год
                summ_ved_duty = summ_ved_duty if summ_ved_duty is not None else 0
                summ_ved_duty_defrayal = self.ved_sheet_list.cell(row=row,
                                                                  column=9).value  # оплата долга за прошлый год
                summ_ved_duty_defrayal = summ_ved_duty_defrayal if summ_ved_duty_defrayal is not None else 0

                summ_difference = summ_ved_duty - summ_ved_duty_defrayal
                # if kv =='12':
                #     print(kv)
                if summ_difference > 0:

                    if self.summ_bank > summ_difference:

                        summ_ved_duty_defrayal += summ_difference
                        self.summ_bank -= summ_difference
                        self.ved_sheet_list.cell(row=row, column=9,
                                                 value=summ_ved_duty_defrayal)  # записываем в ячейку оплаты долга

                        if self.summ_bank > 0:
                            self.record_ved_month(row, summ_ved_fee)

                    elif self.summ_bank <= summ_difference:
                        summ_ved_duty_defrayal += self.summ_bank
                        self.summ_bank -= self.summ_bank
                        self.ved_sheet_list.cell(row=row, column=9,
                                                 value=summ_ved_duty_defrayal)  # записываем в ячейку оплаты долга
                        if self.summ_bank > 0:
                            self.record_ved_month(row, summ_ved_fee)
                    break


                else:
                    self.record_ved_month(row, summ_ved_fee)
                    break
            else:
                pass

    def record_ved_month(self, row, summ_ved_fee):  # Разбиваем сумму с банка по месяцам
        """Разбиваем сумму с банка по месяцам"""
        month_ved_start = ''
        month_ved_end = ''
        for coll in range(10, 23):

            summ_ved = self.ved_sheet_list.cell(row=row, column=coll).value  # сумма за месяц
            summ_ved = summ_ved if summ_ved is not None else 0

            if summ_ved != summ_ved_fee and self.summ_bank > 0:

                month_ved_start = self.ved_sheet_list.cell(row=7, column=coll).value  # название месяца
                summ_ved_bank = self.summ_bank + summ_ved

                while summ_ved_bank > 0:
                    if coll <= 21:

                        if summ_ved_bank >= summ_ved_fee and self.summ_bank > 0:

                            self.ved_sheet_list.cell(row=row, column=coll, value=summ_ved_fee)
                            month_ved_end = self.ved_sheet_list.cell(row=7, column=coll).value
                            summ_ved_bank -= summ_ved_fee
                            self.summ_bank = 0
                            coll += 1

                        elif summ_ved_bank >= summ_ved_fee:

                            self.ved_sheet_list.cell(row=row, column=coll, value=summ_ved_fee)
                            month_ved_end = self.ved_sheet_list.cell(row=7, column=coll).value
                            summ_ved_bank -= summ_ved_fee
                            self.summ_bank = 0
                            coll += 1

                        else:
                            self.ved_sheet_list.cell(row=row, column=coll, value=summ_ved_bank)
                            month_ved_end = self.ved_sheet_list.cell(row=7, column=coll).value
                            summ_ved_bank -= summ_ved_bank
                            self.summ_bank = 0
                            coll += 1

                    elif coll > 21:
                        sum_jan = self.ved_sheet_list.cell(row=row, column=coll).value
                        sum_jan = sum_jan if sum_jan is not None else 0

                        self.ved_sheet_list.cell(row=row, column=coll, value=summ_ved_bank + sum_jan)
                        month_ved_end = self.ved_sheet_list.cell(row=7, column=coll).value
                        self.summ_bank = 0
                        summ_ved_bank -= summ_ved_bank
                else:
                    break
            else:
                pass


        payments_period = self.ved_sheet_payment.cell(row=self.payments_row, column=3).value
        payments_period = payments_period if payments_period is not None else 0

        if payments_period == 0:
            if month_ved_start != month_ved_end:
                month_range = f'{month_ved_start} - {month_ved_end}'
                self.ved_sheet_payment.cell(row=self.payments_row, column=3, value=month_range)
            elif month_ved_start == '':
                pass
            else:
                self.ved_sheet_payment.cell(row=self.payments_row, column=3, value=month_ved_start)
        else:
            month_range = f'{payments_period} - {month_ved_end}'
            self.ved_sheet_payment.cell(row=self.payments_row, column=3, value=month_range)

    def comparison(self):
        """Проверяем суммы из файла банка и оплаты"""
        wb_ved = op.load_workbook(self.path_ved, read_only=True, )  # открываем фаил с ведомостью
        ved_sheet_payment = wb_ved['оплата']  # открываем лист с оплато
        summ_payments = 0
        for row in range(self.payments_coord_row + 2, self.payments_coord_row + 2 + self.kv_quantity):
            s_pa = ved_sheet_payment.cell(row=row, column=4).value
            s_pa = s_pa if s_pa is not None else 0
            summ_payments += s_pa
        summ_bank = self.bank_sheet.cell(row=2, column=4).value
        summ_comparison = summ_bank - summ_payments

        if summ_comparison == 0:
            print('Оплаты всех квартир занесены верно\n')
        elif summ_comparison > 0:
            print(f'В листе "оплата" не хватает суммы {summ_comparison}\n')
        elif summ_comparison < 0:
            print(f'В листе "оплата" сумма больше на {summ_comparison * (-1)}\n')

        self.wb_ved.close()

    def launch(self):
        """Запуск асистента (чтение и запись)"""

        self.bank_reading()
        self.month_coordin()
        self.record_payments()

        try:
            self.wb_ved.save(self.path_ved)
            self.wb_ved.close()
            self.comparison()
            print('Программа завершила работу \n')
        except PermissionError:
            print(f'Ошибка! Вы не закрыли файл "{self.name_ved}"\n')
            print('Программа завершила работу с ошибками\n')


if __name__ == '__main__':
    a = 'D:/PyCharm/Project/OSI/fail/вед  Кокжал Барака.xlsx'
    a1 = 'D:/PyCharm/Project/osi_doc/2/вед  Кокжал Барака 2024г — копия.xlsx'
    b = 'D:/PyCharm/Project/osi_doc/2/Платежи (80).xlsx'
    r = 'D:/PyCharm/Project/OSI/fail/Платежи август банк версия.xlsx'
    # ved('fail/вед  Кокжал Барака.xlsx', 'fail/Платежи (46).xlsx')
    Assistant(a1, b, 60).launch()
