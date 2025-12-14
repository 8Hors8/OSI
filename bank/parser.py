"""
parser.py

Модуль отвечает за загрузку и первичную обработку банковских Excel-файлов.

Функциональность модуля:
- загрузка банковского Excel-файла;
- нормализация даты платежа;
- извлечение номера квартиры из строки назначения платежа;
- валидация данных платежа;
- агрегация платежей по дате и типу;
- формирование структурированных данных для последующей разноски.

Модуль не выполняет запись данных в ведомости и не содержит бизнес-логики ОСИ.
"""

import logging

import openpyxl as op
import re
from openpyxl.workbook import Workbook
from typing import Optional
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

logger = logging.getLogger(__name__)


def load_bank_file(file_path: str) -> Optional[Workbook]:
    """
    Загружает Excel-файл банковской ведомости.

    Открывает файл Excel по указанному пути и возвращает активный лист.
    В случае ошибки чтения логирует причину и возвращает None.

    :param file_path: Полный путь к файлу банковской ведомости.
    :return: Активный лист Excel (Worksheet) или None при ошибке загрузки.
    """
    try:
        # data_only=True — если нужен результат формул
        book = op.load_workbook(file_path, data_only=True)
        sheet = book.active
        logger.info(f'Фйл с банка успешно загружен!')
        return sheet
    except FileNotFoundError:
        logger.error(f'Ошибка: Файл не найден по пути: "{file_path}"')
        return None
    except PermissionError:
        logger.error(f'Ошибка: Нет доступа. Закройте файл "{file_path}"')
        return None
    except InvalidFileException:
        logger.error(f'Ошибка: Файл "{file_path}" не является корректным Excel-файлом.')
        return None
    except Exception as e:
        logger.error(f'Непредвиденная ошибка при загрузке файла: {e}')
        return None


def extract_apartment_number(apartment_data: str) -> Optional[str]:
    """
    Извлекает номер квартиры из строки назначения платежа.

    Ожидается, что строка имеет формат, содержащий номер квартиры
    в шестом элементе, разделённом символом ';'.

    :param apartment_data: Строка с описанием назначения платежа.
    :return: Номер квартиры в виде строки или None при ошибке извлечения.
    """
    apartment_number = apartment_data.split(';')[5]

    return apartment_number


def group_daily_payments(result_payment: list, add_dict: dict) -> Optional[list[dict[str, str | int]]]:
    """
    Группирует платежи по типу счёта и дате платежа.

    Если в списке уже существует платёж с тем же типом счёта
    и датой, сумма платежа увеличивается.
    В противном случае платёж добавляется как новый элемент.

    :param result_payment: Список ранее обработанных платежей.
    :param add_dict: Новый платёж в виде словаря
                     {'type': str, 'sum': int, 'date': str}.
    :return: Обновлённый список платежей.
    """
    typing_payment = add_dict['type']
    for payment in result_payment:
        if payment['type'] == typing_payment and payment['date'] == add_dict['date']:
            payment['sum'] += add_dict['sum']
            return result_payment

    result_payment.append(add_dict)
    return result_payment


def normalize_date(date_obj, ) -> Optional[str]:
    """
    Приводит дату платежа к строковому формату 'ДД.ММ.ГГГГ'.

    Поддерживает:
    - datetime.datetime;
    - строку вида 'YYYY-MM-DD', 'DD.MM.YYYY', 'YYYY-MM-DD HH:MM:SS'.

    :param date_obj: Объект datetime или строка с датой.
    :return: Дата в формате 'ДД.ММ.ГГГГ' или None, если формат неизвестен.
    """

    if isinstance(date_obj, datetime):
        return str(date_obj.strftime('%d.%m.%Y'))

    if isinstance(date_obj, str):
        date_obj = date_obj.strip().split(' ')[0].replace('-', '.')
        return str(date_obj)

    return None


def has_payment_errors(apartment_number: str, sum_amount: int, date_amount: str) -> bool:
    """
    Проверяет корректность данных платежа.

    Функция проверяет наличие None-значений в ключевых полях платежа.
    При обнаружении ошибки логирует соответствующее сообщение.

    :param apartment_number: Номер квартиры.
    :param sum_amount: Сумма платежа.
    :param date_amount: Дата платежа.
    :return: True, если обнаружена хотя бы одна ошибка, иначе False.
    """
    errors = {
        'Квартира': apartment_number,
        'Сумма': sum_amount,
        'Дата': date_amount,
    }

    has_error = False
    for field, value in errors.items():
        if value is None:
            logger.error(f'{field} задан(а) некорректно')
            has_error = True

    return has_error


def acquisition_data(sheet) -> Optional[dict[str, list[dict[str, str]]]]:
    """
    Извлекает и агрегирует данные платежей из банковского Excel-листа.

    Для каждой строки листа:
    - извлекает тип счёта, номер квартиры, сумму и дату платежа;
    - нормализует дату;
    - выполняет валидацию данных;
    - группирует платежи по квартире, типу счёта и дате.

    Итоговая структура данных:
    {
        '12': [
            {'type': 'Текущий счёт', 'sum': 1500, 'date': '01.03.2025'},
            {'type': 'Накопительный счёт', 'sum': 500, 'date': '05.03.2025'}
        ],
        '13': [...]
    }

    :param sheet: Активный лист Excel с банковскими данными.
    :return: Словарь с агрегированными платежами или None при ошибке.
    """
    result = {}
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        if sheet.cell(row=row, column=2).value is not None:

            payment_type = sheet.cell(row=row, column=1).value
            apartment_number = extract_apartment_number(sheet.cell(row=row, column=2).value)
            sum_amount = int(sheet.cell(row=row, column=4).value)
            date_amount = normalize_date(sheet.cell(row=row, column=5).value)

            validator_value = has_payment_errors(apartment_number, sum_amount, date_amount)
            if validator_value:
                logger.error(f'Ошибка в строке {row}')
                continue

            add_dict = {'type': payment_type, 'sum': sum_amount, 'date': date_amount}

            if apartment_number in result:
                result[apartment_number] = group_daily_payments(result[apartment_number], add_dict)
            else:
                result[f'{apartment_number}'] = [add_dict]

    return result


if __name__ == '__main__':
    if not logger.hasHandlers():
        logging.basicConfig(
            level=logging.DEBUG,
            format="[%(asctime)s.%(msecs)03d] %(module)s:%(lineno)d %(levelname)7s - %(message)s"
        )
