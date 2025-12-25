"""
bank_parser.py

Модуль отвечает за загрузку и первичную обработку банковских Excel-файлов.

Функциональность модуля:
- загрузка банковского Excel-файла;
- нормализация даты платежа;
- извлечение номера квартиры из строки назначения платежа;
- валидация данных платежа;
- агрегация платежей по дате и типу;
- формирование структурированных данных для последующей разноски.

Модуль не выполняет запись данных в ведомости и не содержит бизнес-логики ОСИ.
"""

import logging

import openpyxl as op
import re
from openpyxl.workbook import Workbook
from typing import Optional
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime

logger = logging.getLogger(__name__)


def extract_apartment_number(apartment_data: str) -> Optional[str]:
    """
    Извлекает номер квартиры из строки назначения платежа.

    Из шестого элемента строки, разделённой ';',
    удаляет все символы, кроме цифр.

    :param apartment_data: Строка с описанием назначения платежа.
    :return: Номер квартиры (только цифры) или None, если извлечь невозможно.
    """
    if not apartment_data:
        return None

    parts = apartment_data.split(';')
    if len(parts) < 6:
        return None


    raw_value = parts[5]

    # Оставляем ТОЛЬКО цифры
    digits = re.sub(r"\D", "", raw_value)

    return digits if digits else None


def group_daily_payments(result_payment: list, add_dict: dict) -> Optional[list[dict[str, str | int]]]:
    """
    Группирует платежи по типу счёта и дате платежа.

    Если в списке уже существует платёж с тем же типом счёта
    и датой, сумма платежа увеличивается.
    В противном случае платёж добавляется как новый элемент.

    :param result_payment: Список ранее обработанных платежей.
    :param add_dict: Новый платёж в виде словаря
                     {'type': str, 'sum': int, 'date': str}.
    :return: Обновлённый список платежей.
    """
    typing_payment = add_dict['type']
    for payment in result_payment:
        if payment['type'] == typing_payment and payment['date'] == add_dict['date']:
            payment['sum'] += add_dict['sum']
            return result_payment

    result_payment.append(add_dict)
    return result_payment


def normalize_date(date_obj, ) -> Optional[str]:
    """
    Приводит дату платежа к строковому формату 'ДД.ММ.ГГГГ'.

    Поддерживает:
    - datetime.datetime;
    - строку вида 'YYYY-MM-DD', 'DD.MM.YYYY', 'YYYY-MM-DD HH:MM:SS'.

    :param date_obj: Объект datetime или строка с датой.
    :return: Дата в формате 'ДД.ММ.ГГГГ' или None, если формат неизвестен.
    """

    if isinstance(date_obj, datetime):
        return str(date_obj.strftime('%d.%m.%Y'))

    if isinstance(date_obj, str):
        date_obj = date_obj.strip().split(' ')[0].replace('-', '.')
        return str(date_obj)

    return None


def has_payment_errors(apartment_number: str, sum_payment: int, date_payment: str,
                       apartment_number_reference: set) -> tuple[bool, list[dict[str, str | int]]]:
    """
    Проверяет корректность данных платежа.

    Функция проверяет наличие None-значений в ключевых полях платежа.
    При обнаружении ошибки логирует соответствующее сообщение.

    :param apartment_number: Номер квартиры.
    :param sum_payment: Сумма платежа.
    :param date_payment: Дата платежа.
    :param apartment_number_reference: Эталонный набор номеров квартир.
    :return: True, если обнаружена хотя бы одна ошибка, иначе False.
    """
    has_error = False
    log_messages = []

    # квартира
    if apartment_number is None:
        log_messages.append({
            'message': 'Квартира задана некорректно',
            'code': 'INVALID_APARTMENT'
        })
        has_error = True
    elif int(apartment_number) not in apartment_number_reference:
        log_messages.append({
            'message': f'Не корректный номер квартиры "{apartment_number}"',
            'code': 'INVALID_APARTMENT'
        })
        has_error = True

    # сумма
    if sum_payment is None:
        log_messages.append({
            'message': 'Сумма задана некорректно',
            'code': 'INVALID_SUM'
        })
        has_error = True

    # дата
    if date_payment is None:
        log_messages.append({
            'message': 'Дата задана некорректно',
            'code': 'INVALID_DATE'
        })
        has_error = True

    return has_error, log_messages


def acquisition_data(sheet, apartment_number_reference: set) -> Optional[dict[str, list[dict[str, str]]]]:
    """
    Извлекает и агрегирует данные платежей из банковского Excel-листа.

    Для каждой строки листа:
    - извлекает тип счёта, номер квартиры, сумму и дату платежа;
    - нормализует дату;
    - выполняет валидацию данных;
    - группирует платежи по квартире, типу счёта и дате.

    Итоговая структура данных:
    {
        '12': [
            {'type': 'Текущий счёт', 'sum': 1500, 'date': '01.03.2025'},
            {'type': 'Накопительный счёт', 'sum': 500, 'date': '05.03.2025'}
        ],
        '13': [...]
    }

    :param sheet: Активный лист Excel с банковскими данными.
    :param apartment_number_reference: Эталонный набор номеров квартир.
    :return: Словарь с агрегированными платежами или None при ошибке.
    """
    result = {}
    max_row = sheet.max_row
    for row in range(2, max_row + 1):
        if sheet.cell(row=row, column=2).value is not None:

            payment_type = sheet.cell(row=row, column=1).value
            apartment_number = extract_apartment_number(sheet.cell(row=row, column=2).value)
            sum_payment = int(sheet.cell(row=row, column=4).value)
            date_payment = normalize_date(sheet.cell(row=row, column=5).value)

            validator_value, log_messages = has_payment_errors(apartment_number, sum_payment, date_payment,
                                                 apartment_number_reference)
            if validator_value:
                for log in log_messages:
                    logger.error(f'Ошибка в строке {row}: "{log["message"]}',
                    extra={'code':log['code']})
                continue

            add_dict = {'type': payment_type.lower(), 'sum': sum_payment, 'date': date_payment}

            if apartment_number in result:
                result[apartment_number] = group_daily_payments(result[apartment_number], add_dict)
            else:
                result[f'{apartment_number}'] = [add_dict]

    return result


if __name__ == '__main__':
    if not logger.hasHandlers():
        logging.basicConfig(
            level=logging.DEBUG,
            format="[%(asctime)s.%(msecs)03d] %(module)s:%(lineno)d %(levelname)7s - %(message)s"
        )
