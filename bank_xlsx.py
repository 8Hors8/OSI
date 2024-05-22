import openpyxl as op
import re
import pprint as pp


def payments(bank_sheet, name_bank='None'):
    """

    :param bank_sheet:
    :param name_bank:
    :return:
    """
    name_fil = name_bank
    # wb = op.load_workbook(a, data_only=True)
    sheet = bank_sheet

    max_row = sheet.max_row

    # print(f'max_row {max_row}')

    kv_su = {}
    kv_repeat = {}
    for s in range(3, max_row + 1):

        try:
            st_col2 = (sheet.cell(row=s, column=2).value).split(';')
            st_col4 = sheet.cell(row=s, column=4).value
            st_col5 = sheet.cell(row=s, column=5).value

        except AttributeError:
            print('Файл от банка выбран не верно')
            break

        kv_chek = ''.join(re.findall('\d', st_col2[5]))

        if len(kv_chek) <= 3:
            if kv_chek[0] == '0':

                kv = kv_chek[1]

            else:
                kv = kv_chek
        else:
            kv = kv_chek
            print(f'\nОШИБКА № {kv_chek} кв  в файле "{name_fil}" указано не верно, координаты '
                  f'ошибки (строка {s} столбец 2)\n')

        su = st_col4
        date = (st_col5.split(' ')[0]).replace('-', '.')
        n = 2

        def check_date(kv):
            nonlocal su
            nonlocal n
            if kv in kv_su.keys():

                if kv not in kv_repeat.keys() and len(kv) <= 2:

                    kv_repeat[kv] = 2

                elif kv in kv_repeat.keys():

                    values = kv_repeat[kv]
                    kv_repeat[kv] = values + 1

                if kv_su[kv][1] != date:

                    kv_n = kv

                    while kv_n in kv_su.keys():
                        if len(kv_n)<=2:

                            kv_n = f'{kv}-{n}'
                        else:
                            kv_n = kv_n.split('-')
                            kv_n[-1] = str(n)
                            kv_n = '-'.join(kv_n)
                        n += 1
                        kv = kv_n
                        return check_date(kv)


                else:
                    sum_kv = kv_su[kv][0]
                    # date = f'{kv_su[kv][1]}/{date}'
                    su = su + sum_kv

                kv_su[kv] = (su, date)

            else:
                kv_su[kv] = (su, date)

        check_date(kv)

    # pp.pprint((kv_su,'len = ',len(kv_su)))
    if kv_repeat != {}:
        print(f'Квартиры которые встречаются в файле более одного раза: \n{kv_repeat}\n')
    else:
        print('Повторяющийся квартир нет \n')
        pass
    # print('\n',kv_su)
    return kv_su


if __name__ == '__main__':
    b = 'fail/Платежи (46).xlsx'
    y = 'fail/4Платежи апрель банк версия test.xlsx'
    a = 'D:/PyCharm/Project/OSI/fail/Платежи сентябрь бак версия.xlsx'
    c = 'D:/PyCharm/Project/OSI/fail/4Платежи апрель банк версия.xlsx'
    b = 'D:/PyCharm/Project/osi_doc/2/Платежи (80).xlsx'
    p = 'C:/Users/бук/Desktop/fail1/4Платежи апрель банк версия.xlsx.xlsx'
    wb = op.load_workbook(b, data_only=True)
    bank_sheet = wb.active
    t = payments(bank_sheet)
    print(t, 'len-', len(t))
