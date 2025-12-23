"""
statements_parser.py

Модуль содержит инструменты для извлечения данных из Excel-ведомостей
на основе описательных схем.

Основная идея:
- структура Excel не зашита в код;
- каждая таблица описывается схемой (schema);
- парсеры читают данные строго по этим схемам.

Модуль НЕ отвечает за:
- загрузку файлов;
- бизнес-логику;
- агрегацию или сохранение данных.

Назначение модуля — инфраструктурный парсинг Excel.
"""


import logging
from openpyxl import Workbook

logger = logging.getLogger(__name__)


class UniversalScan:
    """
    Универсальный сканер Excel-листов по описательной схеме.

    Класс извлекает данные из Excel-листа на основании схемы, которая
    описывает структуру данных, но не содержит логики их обработки.

    Схема должна определять:
    - NAME_SHEET      — имя листа;
    - ROW_START       — строка заголовка;
    - COLUMN_START    — колонка заголовка;
    - EXPECTED_VALUE  — ожидаемый текст заголовка;
    - SCAN_TYPE       — тип сканирования: "row", "column", "mixed";
    - ROW_OFFSET      — смещение начала данных по строкам;
    - COLUMN_OFFSET   — смещение начала данных по колонкам.

    Класс:
    - не загружает файлы;
    - не хранит состояние приложения;
    - не знает бизнес-правил.

    Его задача — безопасно извлечь данные и отфильтровать невалидные значения.
    """

    def __init__(self, book: Workbook, schema: type):
        """
        Инициализирует сканер.

        :param book: Загруженная Excel-книга (openpyxl.Workbook)
        :param schema: Класс-схема, описывающий структуру листа
        """
        self.schema = schema

        self.sheet = book[getattr(schema, "NAME_SHEET")]
        self.row_start = getattr(schema, "ROW_START")
        self.column_start = getattr(schema, "COLUMN_START")
        self.row_offset = getattr(schema, "ROW_OFFSET", 0)
        self.column_offset = getattr(schema, "COLUMN_OFFSET", 0)
        self.scan_type = getattr(schema, "SCAN_TYPE")

        self.expected_value = str(
            getattr(schema, "EXPECTED_VALUE", "")
        ).lower()

        header_value = self.sheet.cell(
            row=self.row_start,
            column=self.column_start
        ).value

        self.start_cell = (
            header_value.lower()
            if isinstance(header_value, str)
            else ""
        )

    def scan(self) -> list:
        """
        Запускает сканирование листа согласно схеме.

        Сначала проверяет корректность заголовка,
        затем выбирает нужный режим сканирования.

        :return: Список валидных значений
        """
        if self.expected_value not in self.start_cell:
            logger.error(
                f"Ошибка структуры листа: в ячейке "
                f"{self.row_start}:{self.column_start} "
                f"ожидалось '{self.expected_value}', найдено '{self.start_cell}'"
            )
            return []

        logger.debug(f'Заголовок подтвержден: "{self.expected_value}"')

        if self.scan_type == "row":
            return self._row_scan()

        if self.scan_type == "column":
            return self._column_scan()

        if self.scan_type == "mixed":
            return self._mixed_scan()

        logger.error(f"Неизвестный тип сканирования: {self.scan_type}")
        return []

    def _row_scan(self) -> list:
        """
        Сканирует значения вниз по одному столбцу.

        :return: Список значений
        """
        result = []
        column = self.column_start + self.column_offset

        for row in range(
            self.row_start + self.row_offset,
            self.sheet.max_row + 1
        ):
            value = self.sheet.cell(row=row, column=column).value

            if self._validate_value(value):
                result.append(value)
            else:
                logger.warning(
                    f"Невалидное значение в ячейке {row}:{column} — {value}"
                )

        return result

    def _column_scan(self) -> list:
        """
        Сканирует значения вправо по одной строке.

        :return: Список значений
        """
        result = []
        row = self.row_start + self.row_offset

        for column in range(
            self.column_start + self.column_offset,
            self.sheet.max_column + 1
        ):
            value = self.sheet.cell(row=row, column=column).value

            if self._validate_value(value):
                result.append(value)
            else:
                logger.warning(
                    f"Невалидное значение в ячейке {row}:{column} — {value}"
                )

        return result

    def _mixed_scan(self) -> list:
        """
        Сканирует значения в табличной структуре (строки и столбцы).

        :return: Список значений
        """
        result = []

        for row in range(
            self.row_start + self.row_offset,
            self.sheet.max_row + 1
        ):
            for column in range(
                self.column_start + self.column_offset,
                self.sheet.max_column + 1
            ):
                value = self.sheet.cell(row=row, column=column).value

                if self._validate_value(value):
                    result.append(value)
                else:
                    logger.warning(
                        f"Невалидное значение в ячейке {row}:{column} — {value}"
                    )

        return result

    def _validate_value(self, value) -> bool:
        """
        Проверяет одно значение по правилам схемы.

        Метод является инфраструктурным валидатором и не выбрасывает исключений.

        :param value: Значение из ячейки Excel
        :return: True — значение валидно, False — игнорируется
        """

        # 1. Игнорируемые значения
        ignore_values = getattr(self.schema, "IGNORE_VALUES", set())
        if value in ignore_values:
            return False

        # 2. Пустые значения
        allow_empty = getattr(self.schema, "ALLOW_EMPTY", True)
        if value is None:
            return allow_empty

        # 3. Тип значения
        value_type = getattr(self.schema, "VALUE_TYPE", None)
        if value_type is not None and not isinstance(value, value_type):
            return False

        # 4. Диапазон значений
        min_value = getattr(self.schema, "MIN_VALUE", None)
        max_value = getattr(self.schema, "MAX_VALUE", None)

        if min_value is not None and value < min_value:
            return False

        if max_value is not None and value > max_value:
            return False

        return True

